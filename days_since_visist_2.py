import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from datetime import date, datetime


class App():
    

    def get_stores(self):
        """
        Imports an xslx file containing all stores for a given region
        """
        file_path = filedialog.askopenfilename()
        wb1 = load_workbook(filename=file_path)
        store_sheet = wb1[wb1.sheetnames[0]]
        if store_sheet:
            storeLabel = tk.Label(text="Butikslista importerad")
            storeLabel.place(x=160, y=10)
            
        self.store_sheet = store_sheet


    def get_visists(self):
        """
        Imports the excel file containing visits and their dates
        """
        file_path = filedialog.askopenfilename()
        wb2 = load_workbook(filename=file_path)
        visit_sheet = wb2[wb2.sheetnames[0]]
        if visit_sheet:
            visitLabel = tk.Label(text="Besökslista importerad")
            visitLabel.place(x=160, y=40)
            
        self.visit_sheet = visit_sheet
        
    
    def get_stock(self):
        """
        Imports an optional excel file containing the stock for each store
        """
        file_path = filedialog.askopenfilename()
        wb3 = load_workbook(filename=file_path)
        stock_sheet = wb3[wb3.sheetnames[0]]
        if stock_sheet:
            stockLabel = tk.Label(text="Kort i Butik importerad")
            stockLabel.place(x=180, y=70)

        self.stock_sheet = stock_sheet

    def export(self):
        """
        Work through the two imported worksheets,
        to create a new excel file with store and days since you visisted. 
        """
        # creating dictionary on form {butik: {adress: , ort: , Tid: , Lager: }}
        region_dict = {row[0].value: {"Adress": row[1].value, "Ort": row[2].value, "Tid": "N/A", "Lager": "Not Found"}
                    for row in self.store_sheet}

        # iterate over visit_sheet to extract visit date
        for row in self.visit_sheet:
            if row[0].value in region_dict:
                # date.today() - i[2].value.date() is a timedelta, used .days as days object was wanted. 
                if region_dict[row[0].value]["Tid"] == "N/A" or region_dict[
                        row[0].value]["Tid"] > (date.today() -
                                            row[2].value.date()).days:
                    # The result is in hours, days wanted, so .days is used.
                    if isinstance(row[2].value, datetime):
                        region_dict[row[0].value]["Tid"] = (date.today() -
                                                        row[2].value.date()).days
                        
        if hasattr(self, 'stock_sheet'):
            for row in self.stock_sheet:
                if row[0].value in region_dict:
                    region_dict[row[0].value]["Lager"] = row[1].value

        # Creating new workbook and new active worksheet. 
        export = Workbook()
        export_sheet = export.active

    
        
        # Write to file
        # store is the key to the inner dictionary 
        for store in region_dict:
            export_sheet.append([
                store, region_dict[store]["Adress"], region_dict[store]["Ort"],
                region_dict[store]["Tid"], region_dict[store]["Lager"]
            ])

        file = filedialog.asksaveasfile(defaultextension=".xlsx")
        export.save(file.name)


# Initalize class object.
store_app = App()

root = tk.Tk()
root.title("Dagar sedan besök")
root.geometry("500x200+500+500")
root.configure(bg='lightblue')

importStores = tk.Button(root,
                         text="Importera Regionslista",
                         command=store_app.get_stores,
                         bg="lightgrey")
importStores.place(x=10, y=10)
importStores.configure(border=2, relief="raised")

importVisits = tk.Button(root,
                         text="Importera besökslista  ",
                         command=store_app.get_visists,
                         bg="lightgrey")
importVisits.place(x=10, y=40)
importVisits.configure(border=2, relief="raised")

importStock = tk.Button(root,
                        text="Valfri: Importera lagervärde",
                        command=store_app.get_stock,
                        bg="lightgrey")
importStock.place(x=10, y=70)
importStock.configure(border=2, relief="raised")

exportButton = tk.Button(root,
                         text="Exportera besökslista  ",
                         command=store_app.export,
                         bg="orange")
exportButton.place(x=10, y=160)
exportButton.configure(border=2, relief="raised")

quitButton = tk.Button(root, text="Avsluta", command=root.quit)
quitButton.place(x=430, y=160)
quitButton.configure(border=2, relief="raised")

root.mainloop()
